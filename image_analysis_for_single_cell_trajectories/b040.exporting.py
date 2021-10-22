# -*- coding: utf-8 -*-
"""
export excel
"""
import os
from skimage import io
import numpy as np
from scipy import ndimage as ndimage
from skimage import measure
import pickle
import openpyxl



def find_bwc(im,bw):
    bw_high = im * bw > 20
    bw_high2 = ndimage.binary_fill_holes(bw_high).astype(bool)
    seg_high = measure.label(bw_high2)
    if np.max(seg_high) == 1:
        bwc = seg_high == 1
    elif np.max(seg_high) == 0:
        bwc = bw_high2.copy()
    else:
        max_size = 0
        max_label = 0
        for label in range(1,np.max(seg_high)):
            area = np.sum(seg_high == label)
            if area > max_size:
                max_size = area
                max_label = label
        bwc = seg_high == max_label
    return bwc

#%%
path = './'
seg2_file = []
tifs = []
for folderName, subfolders, filenames in os.walk(path):
    for filename in filenames:
        if 'pkl' in filename:
            seg2_file.append(folderName+filename)
        if 'tif' in filename:
            tifs.append(folderName+filename)
        

#%%
frame_starting = 0

tif_index = 0
    
im = io.imread(tifs[tif_index])

segfile = seg2_file[tif_index]
f = open(segfile,'rb')
seg2s = pickle.load(f)
f.close()

excelfile = seg2_file[tif_index][:-13] + '.xlsx'


sheet_name = ['cell size no nuc',
               'cell S mean no nuc',
               'cell S median no nuc',
               'nuc size',
               'nuc mean S',
               'nuc median S',
               'IF']
col_names = []
col_names.append('cell label')
for i in range(len(seg2s)):
    col_names.append('frame'+str(i).zfill(2))

wb = openpyxl.Workbook()
wb.get_sheet_by_name
for sheet_index in range(len(sheet_name)):
    if sheet_index == 0:
        sheet = wb['Sheet']
        sheet.title = sheet_name[sheet_index]
    else:
        sheet = wb.create_sheet(sheet_name[sheet_index])
    for i in range(len(col_names)):
        sheet.cell(row= 1,column=i+1,value=col_names[i])
wb.save(excelfile)

#%%
wb = openpyxl.load_workbook(excelfile)
row_index = 2

labels = list(np.unique(seg2s[0]))
if 0 in labels:
    labels.remove(0)

print(len(labels)-1)    
for target_index in range(len(labels)):
    target = labels[target_index]
    if target_index % 10 == 0:
        print(str(target_index).zfill(3))
    else:
        print(str(target_index).zfill(3),end=', ')
    # for first column
    for sheet_index in range(len(sheet_name)):
        sheet = wb[sheet_name[sheet_index]]
        sheet.cell(row=target_index+1,column=1,value='label.'+str(target).zfill(3))
    
    # for 2nd col and so on
    for frame_index in range(len(seg2s)):
        label_current = list(np.unique(seg2s[frame_index]))
        if target in label_current:

            bw = seg2s[frame_index] == target
            img = im[frame_index]
            bwc = find_bwc(img,bw)
            bwo = bw ^ bwc
            output = []
            output.append(np.sum(bwo))

            if np.sum(bwo) == 0:
                output.append(0)
                output.append(0)
            else:
                output.append(np.mean(img[bwo]))
                output.append(np.median(img[bwo]))

            output.append(np.sum(bwc))
    
            if np.sum(bwc) == 0:
                output.append(0)
                output.append(0)
            else:
                output.append( np.mean(img[bwc]) )
                output.append(np.median(img[bwc]))
    
            if output[3] < 15:
                output.append(output[1])
            else:
                output.append(output[4])
        else:
            output = [0,0,0,0,0,0,0]
    
                
        for sheet_index in range(len(sheet_name)):
            sheet = wb[sheet_name[sheet_index]]
            sheet.cell(row=target_index+1,column=frame_index +2,
                           value=output[sheet_index])
        


wb.save(excelfile)
