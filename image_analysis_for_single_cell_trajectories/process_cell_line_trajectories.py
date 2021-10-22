#!/usr/bin/env python
# coding: utf-8

# In[1]:


import glob
import openpyxl 
import numpy as np
import matplotlib.pyplot as plt
import os
from scipy.signal import find_peaks
import seaborn as sns
import pandas as pd

plt.rcParams['font.family'] = 'Arial'


# In[2]:


def cm_to_inch(value):
    return value/2.54


# In[3]:


def copy_columns(source_ws, destination_ws, col0, col1):
    for cell in source_ws[col0+":"+col1]:
        destination_ws.cell(row=cell.row, column=cell.column, value=cell.value)


# In[4]:


def copy_rows(source_ws, destination_ws, row0, row1):
    for cell in source_ws[row0:row1]:
        destination_ws.cell(row=cell.row, column=cell.column, value=cell.value)


# In[5]:


def low_pass_filter(t,s,fc=10**32,plot=False,xlim=None, plot_s=False, shape='gaussian'):
    fourierTransform  = np.fft.fft(s)
    samplingTimeStep = t[-1]-t[-2]
    FrequencyRange = 1 / samplingTimeStep
    tpCount     = len(s)
    values      = np.arange(int(tpCount))
    values[-1:-tpCount//2:-1]=np.arange(-1,-tpCount//2,-1)
    timePeriod  = tpCount/FrequencyRange
    frequencies = values/timePeriod
    
    if plot:
        plt.plot(frequencies, abs(fourierTransform)/len(s))
        if xlim:
            plt.xlim(*xlim)
        plt.show()
    
    if shape=='square':
        square=abs(frequencies)<=fc
        fourierTransform *= square
    if shape=='gaussian':
        factor=np.exp(-(frequencies)**2/2/fc**2)
        fourierTransform *= factor
    if shape=='sinc':
        factor=np.sinc((frequencies)/fc)
        fourierTransform *= factor

    itx = np.fft.ifft(fourierTransform)
    if plot_s:
        plt.plot(t,itx.real,label='filtered')
        plt.plot(t,s,label='original')
        plt.show()
        
    return itx.real


# In[6]:


def assign_them(pk,pkph,fft_res):
    
    peak_phase   = np.array([np.nan]*3)
    peak         = np.array([np.nan]*3)
    period       = np.array([np.nan]*2)
    trough       = np.array([np.nan]*2)
    trough_phase = np.array([np.nan]*2)
    amplitude    = np.array([np.nan]*3 )   
        
    for i in range(len(pkph)):
        if pkph[i]<=20:
            peak_phase[0] = pkph[i]
            peak[0] = pk[i]
        elif pkph[i]>20 and pkph[i]<41:
            peak_phase[1] = pkph[i]
            peak[1] = pk[i]            
        else:
            peak_phase[2] = pkph[i]
            peak[2] = pk[i]   
            
    period = peak_phase[1:] - peak_phase[0:-1]
    
    trough_phase = (peak_phase[1:] + peak_phase[0:-1])/2
    for i in range(len(trough_phase)):
        tp = trough_phase[i]
        if ~np.isnan(tp):
            trough[i] = fft_res[int(tp)]
    
    n = len(peak)-1
    for i in range(n):
        amplitude[i] = peak[i] - max(0,trough[i])
        
    amplitude[n] = peak[n] - max(0,trough[n-1])
    return peak,peak_phase,period,trough_phase,trough,amplitude


# In[7]:


def organized_cellfiles():
    files = glob.glob('*cut.xlsx')
    first = set()
    for file in sorted(files):
        print(file)
        wb_obj = openpyxl.load_workbook(file)
        nfile = file.split(".")[0]+".xlsx"
        sub = "_"+file.split(".")[1]

        posi_label = file.split("_")[1].split(".")[0].replace(" ","").lower()

        sdata = []
        cdata = []
        fdata = []

        if nfile not in first:
            first.add(nfile)
            try:
                os.remove(nfile)
            except:
                pass

            sdata.append([
                "labels",
                "peakphase1","peakphase2","peakphase3",
                "peak1","peak2","peak3",
                "troughphase1","troughphase2",
                "trough1","trough2",
                "amplitude1","amplitude2","amplitude3",
                "period1","period2","period3"
            ])

        try:
            wb_out = openpyxl.load_workbook(nfile)
        except:
            wb_out = openpyxl.Workbook()
            try:
                wb_out.remove(wb_out["Sheet"])
            except:
                pass

        nuc_size = wb_obj["nuc size"]
        nuc_mean_S = wb_obj["nuc mean S"]
        cell_S_mean_no_nuc = wb_obj["cell S mean no nuc"]

        ws = "read"
        if ws not in wb_out.sheetnames:
            wb_out.create_sheet(ws)
        current_ws =  wb_out[ws]
        copy_rows(nuc_size, current_ws, 1, 1)
        #copy_columns(nuc_size, current_ws, "A", "A")

        ws = "FFT"
        if ws not in wb_out.sheetnames:
            wb_out.create_sheet(ws)
        FFT_filtered = wb_out[ws]
        copy_rows(nuc_size, FFT_filtered,1, 1)
        #copy_columns(nuc_size, FFT_filtered, "A", "A")

        ws = "SUMMARY_stat"
        if ws not in wb_out.sheetnames:
            wb_out.create_sheet(ws)
        peak_phase_trough_amplitude = wb_out[ws]

        rows = nuc_size["B:B"]
        cols = nuc_size["2:2"]

        cmfft = []

        for rcell in rows:        
            r = rcell.row
            if r:
                clabel = nuc_size.cell(row=r, column=1).value
                s = np.array([],dtype='float')
                for ccell in cols:
                    c = ccell.column
                    if c>1:
                        cond = nuc_size.cell(row=r, column=c).value
                        if cond == None:
                            cond = 0
                        try:
                            v = nuc_mean_S.cell(row=r, column=c).value if cond > 15 else cell_S_mean_no_nuc.cell(row=r, column=c).value
                        except:
                            v = None

                        s = np.append(s,v)

                if None not in s:
                    cdata.append([posi_label+"."+clabel, *s])

                    t = np.array(list(range(len(s))))
                    fft_res = low_pass_filter(t,s,fc=0.05,plot=False,xlim=None, plot_s=False, shape='gaussian') 
                    cmfft.append(fft_res)

                    fdata.append([posi_label+"."+clabel, *fft_res])

                    posi = np.array(find_peaks(fft_res, distance=15,height=10)[0])
                    posi = posi[(posi<len(s)-1) & (posi>0)]
                    pk   = fft_res[posi] 
                    pkph = t[posi]

                    peak,peak_phase,period,trough_phase,trough,amplitude = assign_them(pk,pkph,fft_res)

                    sdata.append([ 
                        posi_label+"."+clabel, 
                        *peak_phase, 
                        *peak, 
                        *trough_phase,
                        *trough,
                        *amplitude,
                        *period  
                    ])

                    #plt.plot(t,fft_res,peak_phase, peak,'xk',trough_phase, trough, '+k')  

        for crow in cdata:
            current_ws.append(crow)

        for frow in fdata:
            FFT_filtered.append(frow)                

        for srow in sdata:
            peak_phase_trough_amplitude.append(srow)

        wb_obj.save(file)
        wb_out.save(nfile)


# In[8]:


def generate_heat_map():
    files = glob.glob('*[!cut].xlsx')
    ffmap = plt.get_cmap()

    sns.set_context("paper",font_scale=1.5)
    col_div = 5
    row_div = 5
    fig, axs = plt.subplots(row_div, col_div,figsize=(15,15))

    ind = 0
    ind2 = -1
    for file in sorted(files):
        print(file)
        wb_obj = openpyxl.load_workbook(file)

        read = wb_obj["read"]
        ffts = wb_obj["FFT"]
        stat = wb_obj["SUMMARY_stat"]

        ws = "plots"
        if ws in wb_obj.sheetnames:
            wb_obj.remove(wb_obj[ws])
        wb_obj.create_sheet(ws)
        ws_plot =  wb_obj[ws]

        rows = read["B:B"]
        cols = read["2:2"]

        cmfft = {}
        cmfft["intensity"] = []
        cmfft["time (hours)"] = []
        cmfft["cell-lines"] = []

        for rcell in rows:        
            r = rcell.row
            if r>1:
                clabel = read.cell(row=r, column=1).value
                s = np.array([],dtype='float')
                for ccell in cols:
                    c = ccell.column
                    if c>1:
                        v = read.cell(row=r, column=c).value
                        #v = ffts.cell(row=r, column=c).value
                        s = np.append(s,v)

                try:
                    #s = (s-np.mean(s)) / np.std(s)
                    #s = s + abs(np.min(s))
                    cmax = np.max(s)
                    s = s / cmax
                    if np.nan != np.sum(s):
                        if cmax>10:
                        #if True:
                            [ cmfft["time (hours)"].append(x) for x in range(len(s)) ]
                            [ cmfft["cell-lines"].append(clabel) for x in s ]
                            [ cmfft["intensity"].append(x) for x in s ]
                except:
                    pass

        df = pd.DataFrame(cmfft)
        table = pd.pivot_table(df, values="intensity", index=["cell-lines"], columns="time (hours)")

        if ind%col_div == 0:
            ind2 = ind2 + 1
        if not table.empty:
            cg = sns.clustermap(table,col_cluster=False,figsize=(10, 10),cmap=plt.get_cmap(),xticklabels=False,yticklabels=False,standard_scale=0)
            #order_col = cg.dendrogram_col.reordered_ind
            order_row = cg.dendrogram_row.reordered_ind
            #print(dir(cg.ax_heatmap))
            plt.close(cg.fig)        

            #table = table.iloc[order_row].reindex(columns = order_col)
            #ffmap =  'Spectral_r'
            table = table.iloc[order_row]
            ax = sns.heatmap(table,cbar=False,xticklabels=False,yticklabels=False,ax=axs[ind2,ind%col_div],cmap=ffmap)  
            ax.set(xlabel=file.replace(".xlsx",""))
            plt.tight_layout()
            ax.set(ylabel="")

        wb_obj.save(file)
        ind = ind + 1
        #break

    fig.set_tight_layout(True)
    fig.delaxes(axs[4][3])
    fig.delaxes(axs[4][4])
    #fig.delaxes(axs[ind2][ind%col_div])
    plt.savefig("heat_map_subplots.png",dpi=300)


# In[9]:


def generate_heat_map2():
    files = glob.glob('*[!cut].xlsx')
    ffmap = plt.get_cmap()

    ind = 0
    ind2 = -1
    for file in sorted(files):
        
        
        sns.set_context("paper") #font_scale=1.5
        plt.rc('xtick', labelsize=8)
        plt.rc('ytick', labelsize=8)
        plt.rcParams['font.sans-serif'] = 'Arial'
        plt.rcParams['font.size'] = 8
        plt.rcParams['axes.linewidth'] = 1.1
        plt.rcParams['mathtext.rm'] = 'Arial'
        plt.figure(figsize=(cm_to_inch(8),cm_to_inch(8)))            
        
        print(file)
        wb_obj = openpyxl.load_workbook(file)

        read = wb_obj["read"]
        #ffts = wb_obj["FFT"]
        #stat = wb_obj["SUMMARY_stat"]

        ws = "plots"
        if ws in wb_obj.sheetnames:
            wb_obj.remove(wb_obj[ws])
        wb_obj.create_sheet(ws)
        ws_plot =  wb_obj[ws]

        rows = read["B:B"]
        cols = read["2:2"]

        cmfft = {}
        cmfft["intensity"] = []
        cmfft["time (hours)"] = []
        cmfft["cell-lines"] = []

        for rcell in rows:        
            r = rcell.row
            if r>1:
                clabel = read.cell(row=r, column=1).value
                s = np.array([],dtype='float')
                for ccell in cols:
                    c = ccell.column
                    if c>1:
                        v = read.cell(row=r, column=c).value
                        #v = ffts.cell(row=r, column=c).value
                        s = np.append(s,v)

                try:
                    #s = (s-np.mean(s)) / np.std(s)
                    #s = s + abs(np.min(s))
                    cmax = np.max(s)
                    s = s / cmax
                    if np.nan != np.sum(s):
                        if cmax>5:
                        #if True:
                            [ cmfft["time (hours)"].append(x) for x in range(len(s)) ]
                            [ cmfft["cell-lines"].append(clabel) for x in s ]
                            [ cmfft["intensity"].append(x) for x in s ]
                except:
                    pass

        df = pd.DataFrame(cmfft)
        table = pd.pivot_table(df, values="intensity", index=["cell-lines"], columns="time (hours)")

        if not table.empty:
            cg = sns.clustermap(table,col_cluster=False,figsize=(cm_to_inch(8), cm_to_inch(8)),cmap=plt.get_cmap(),xticklabels=False,yticklabels=False,standard_scale=0)
            #order_col = cg.dendrogram_col.reordered_ind
            order_row = cg.dendrogram_row.reordered_ind
            #print(dir(cg.ax_heatmap))
            plt.close(cg.fig)        

            #table = table.iloc[order_row].reindex(columns = order_col)
            #ffmap =  'Spectral_r'
            table = table.iloc[order_row]
            ax = sns.heatmap(table,cbar=False,xticklabels=False,yticklabels=False,cmap=ffmap,cbar_kws={"orientation": "horizontal"})  
            ax.set(xlabel=file.replace(".xlsx",""))
            plt.tight_layout()
            ax.set(ylabel="")
            ax.set(xlabel="")

        wb_obj.save(file)
        ind = ind + 1
        
        plt.tight_layout()
        plt.savefig(file+"heat_map_seperated.png",dpi=300)
        #break


# In[10]:


def generte_time_series():
    files = glob.glob('*[!cut].xlsx')
    ffmap = plt.get_cmap()

    plt.rc('xtick', labelsize=8)
    plt.rc('ytick', labelsize=8)
    plt.rcParams['font.sans-serif'] = 'Arial'
    plt.rcParams['font.size'] = 8
    plt.rcParams['axes.linewidth'] = 1.1
    plt.rcParams['mathtext.rm'] = 'Arial'

    col_div = 2
    row_div = 12
    fig, axs = plt.subplots(row_div, col_div,figsize=(15,20), sharex=False, sharey=False)

    ind = 0
    ind2 = -1
    for file in sorted(files):
        print(file)
        wb_obj = openpyxl.load_workbook(file)

        read = wb_obj["read"]
        ffts = wb_obj["FFT"]
        stat = wb_obj["SUMMARY_stat"]

        ws = "plots"
        if ws in wb_obj.sheetnames:
            wb_obj.remove(wb_obj[ws])
        wb_obj.create_sheet(ws)
        ws_plot =  wb_obj[ws]

        rows = read["B:B"]
        cols = read["2:2"]

        cmfft = {}
        cmfft["intensity"] = []
        cmfft["time (hours)"] = []
        cmfft = []

        for rcell in rows:        
            r = rcell.row
            if r>1:
                clabel = read.cell(row=r, column=1).value
                s = np.array([],dtype='float')
                for ccell in cols:
                    c = ccell.column
                    if c>1:
                        v = read.cell(row=r, column=c).value
                        #v = ffts.cell(row=r, column=c).value
                        s = np.append(s,v)

                try:
                    if np.nan != np.sum(s):
                        cmfft.append(s)
                except:
                    pass

        if ind%col_div == 0:
            ind2 = ind2 + 1

        for row in cmfft:
            t = np.arange(0,len(row))
            axs[ind2,ind%col_div].plot(t,row, '-', color='lightgray',lw=0.5)
            #axs[ind2,ind%col_div].set_ylim(bottom=0)

        mean_signal=np.mean(cmfft,axis=0)
        std_signal=np.std(cmfft,axis=0)
        axs[ind2,ind%col_div].errorbar(t,mean_signal,std_signal,fmt='--', zorder=300,lw=1, markeredgewidth=0.6,lolims=True, label=file.replace(".xlsx",""), linestyle='None',capsize=1,marker='.',ms=5)
        axs[ind2,ind%col_div].set_ylim(0,150)
        axs[ind2,ind%col_div].legend(loc ='upper right')
        ind = ind + 1
    fig.set_tight_layout(True)
    fig.delaxes(axs[ind2][ind%col_div])
    plt.savefig("time_series_subplots.png",dpi=300)


# In[11]:


def generte_time_series2():
    files = glob.glob('*[!cut].xlsx')
    ffmap = plt.get_cmap()

    ind = 0
    ind2 = -1
    for file in sorted(files):
        print(file)
        
        sns.set_context("paper") #font_scale=1.5
        plt.rc('xtick', labelsize=8)
        plt.rc('ytick', labelsize=8)
        plt.rcParams['font.sans-serif'] = 'Arial'
        plt.rcParams['font.size'] = 8
        plt.rcParams['axes.linewidth'] = 1.1
        plt.rcParams['mathtext.rm'] = 'Arial'
        plt.figure(figsize=(cm_to_inch(8),cm_to_inch(4)))             
        
        
        wb_obj = openpyxl.load_workbook(file)

        read = wb_obj["read"]
        #ffts = wb_obj["FFT"]
        #stat = wb_obj["SUMMARY_stat"]

        ws = "plots"
        if ws in wb_obj.sheetnames:
            wb_obj.remove(wb_obj[ws])
        wb_obj.create_sheet(ws)
        ws_plot =  wb_obj[ws]

        rows = read["B:B"]
        cols = read["2:2"]

        cmfft = {}
        cmfft["intensity"] = []
        cmfft["time (hours)"] = []
        cmfft = []

        for rcell in rows:        
            r = rcell.row
            if r>1:
                clabel = read.cell(row=r, column=1).value
                s = np.array([],dtype='float')
                for ccell in cols:
                    c = ccell.column
                    if c>1:
                        v = read.cell(row=r, column=c).value
                        #v = ffts.cell(row=r, column=c).value
                        s = np.append(s,v)

                try:
                    if np.nan != np.sum(s):
                        cmfft.append(s)
                except:
                    pass

        for row in cmfft:
            t = np.arange(0,len(row))
            plt.plot(t,row, '-', color='lightgray',lw=0.5)

        mean_signal=np.mean(cmfft,axis=0)
        std_signal=np.std(cmfft,axis=0)
        plt.errorbar(t,mean_signal,std_signal,fmt='--', zorder=300,lw=1, markeredgewidth=0.6,lolims=True, label=file.replace(".xlsx",""), linestyle='None',capsize=1,marker='.',ms=5)
        plt.ylim(0,150)
        plt.xlim(0,61)
        #plt.legend(loc ='upper right')

        plt.tight_layout()
        plt.gca().set_xticks([])
        plt.xticks([])
        plt.savefig(file+"time_series_separated.png",dpi=300)
        plt.close()


# In[12]:


# organized_cellfiles()   # Process cell-lines from excell files by using FFT and collect peaks, trough, amplitude, phases, etc.


# In[13]:


generate_heat_map()     # Generate heap maps of data from previos run into subplots


# In[14]:


generate_heat_map2()    # Generate heap maps of data from previos run into individual plots


# In[15]:


generte_time_series()   # Generate htime series plot of data from previos run into subplots


# In[16]:


generte_time_series2()  # Generate htime series plot of data from previos run into individual plots

