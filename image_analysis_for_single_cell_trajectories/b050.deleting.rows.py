# -*- coding: utf-8 -*-
"""
after export, delete some rows
"""
import numpy as np
import openpyxl

#%% prepare excel file
path = './data/20210917 89h Le/toc1Le5/'
position = 35
excelfile = path + path[path.index('/',7)+1:-1]+'.MAX_Position ' + str(position).zfill(2) +'.xlsx'
sheet_name = ['cell size no nuc',
               'cell S mean no nuc',
               'cell S median no nuc',
               'nuc size',
               'nuc mean S',
               'nuc median S',
               'IF']

#%%
wb = openpyxl.load_workbook(excelfile)
sheet = wb[sheet_name[3]]
row_delete = []
print('original data rows: ',sheet.max_row-1)
for row_index in range(2,sheet.max_row+1):
    row = sheet[row_index]
    values = []
    for col_index in range(1,sheet.max_column):
        values.append(row[col_index].value)

    if np.sum(np.array(values) == 0) > sheet.max_column*2/3:
        row_delete.append(row_index)
    
    if np.sum(np.array(values) > 0) == len(values):
        row_delete.append(row_index)
    
    if np.sum(np.array(values) > 1800) > 0:
        row_delete.append(row_index)

row_delete = np.unique(row_delete)
print('will delete: ',len(row_delete))
print('remains: ', sheet.max_row-1-len(row_delete))
for i in range(len(row_delete)-1,-1,-1):
    target = row_delete[i]
    for j in range(len(sheet_name)):
        sheet = wb[sheet_name[j]]
        sheet.delete_rows(target)

wb.save(excelfile[:-5]+'.cut.xlsx')
